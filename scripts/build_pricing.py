# -*- coding: utf-8 -*-
"""Generate the nexdevp pricing workbook (3 regions + live calculator)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ── Brand palette ──
GREEN = "22B561"
DARK = "0E0E10"
GREY = "8A8A8E"
LIGHT = "F4F4F5"
WHITE = "FFFFFF"

bold_white = Font(name="Calibri", bold=True, color=WHITE, size=11)
title_font = Font(name="Calibri", bold=True, color=DARK, size=18)
sub_font = Font(name="Calibri", italic=True, color="555555", size=10)
hdr_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
green_fill = PatternFill("solid", fgColor=GREEN)
dark_fill = PatternFill("solid", fgColor=DARK)
light_fill = PatternFill("solid", fgColor=LIGHT)
input_fill = PatternFill("solid", fgColor="FFF7CC")
result_fill = PatternFill("solid", fgColor="DFF5E6")
thin = Side(style="thin", color="DDDDDD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
right = Alignment(horizontal="right", vertical="center")

wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════
# SHEET 1 — TARIFAS (3 regiones)
# ════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Tarifas 3 regiones"
ws.sheet_view.showGridLines = False

ws["A1"] = "nexdevp — Estructura de precios"
ws["A1"].font = title_font
ws["A2"] = "Software a medida (proyectos) + Ingeniería de ventas con IA (recurrente). España en EUR · EEUU y LATAM en USD."
ws["A2"].font = sub_font

def header_row(ws, row, cols):
    for i, c in enumerate(cols):
        cell = ws.cell(row=row, column=1 + i, value=c)
        cell.fill = dark_fill if i == 0 else green_fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = border

def data_row(ws, row, vals, money_cols=()):
    for i, v in enumerate(vals):
        cell = ws.cell(row=row, column=1 + i, value=v)
        cell.border = border
        cell.alignment = left if i == 0 else center
        if i == 0:
            cell.font = Font(bold=True, size=10)
        if i % 2 == 1:
            cell.fill = light_fill

def section(ws, row, text):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, color=GREEN, size=13)
    return row + 1

r = 4
r = section(ws, r, "PILAR 1 — Software a medida (precio por proyecto)")
header_row(ws, r, ["Paquete", "España (EUR)", "EEUU (USD)", "LATAM (USD)", "Plazo"])
r += 1
p1 = [
    ["Diagnóstico (se descuenta del proyecto)", "900 – 1.500", "1.500 – 2.500", "500 – 900", "1 sem"],
    ["MVP / Starter", "6.000 – 9.000", "12.000 – 20.000", "3.500 – 6.000", "4–6 sem"],
    ["Sistema a medida (tipo CocinerHosp)", "12.000 – 20.000", "25.000 – 45.000", "8.000 – 15.000", "8–12 sem"],
    ["Enterprise / multi-módulo", "20.000+", "50.000+", "15.000+", "A medida"],
]
for row_vals in p1:
    data_row(ws, r, row_vals); r += 1

r += 1
r = section(ws, r, "PILAR 2 — Ingeniería de ventas con IA (setup + recurrente)")
header_row(ws, r, ["Concepto", "España (EUR)", "EEUU (USD)", "LATAM (USD)", "Tipo"])
r += 1
p2 = [
    ["Setup inicial (una vez)", "1.500 – 3.500", "2.500 – 6.000", "450 – 1.500", "Único"],
    ["Plan Esencial", "149 – 199", "99 – 199", "99 – 149", "Mensual"],
    ["Plan Pro", "299 – 399", "299 – 800", "199 – 299", "Mensual"],
    ["Plan Scale", "599+", "1.000+", "399+", "Mensual"],
]
for row_vals in p2:
    data_row(ws, r, row_vals); r += 1

r += 1
n = ws.cell(row=r, column=1, value="⚠ Importante: el coste de mensajes de WhatsApp (Meta cobra por mensaje desde jul-2025) y de tokens de IA debe estar incluido en el plan o facturado aparte. Definí un volumen incluido y cobrá el excedente.")
n.font = Font(italic=True, color="A05A00", size=9)
n.alignment = left
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 2

r = section(ws, r, "MANTENIMIENTO (Pilar 1) — regla: 15–20% del coste del proyecto al año")
header_row(ws, r, ["Plan", "España (EUR/mes)", "EEUU (USD/mes)", "LATAM (USD/mes)", "Incluye"])
r += 1
mant = [
    ["Básico", "99 – 149", "199 – 399", "59 – 99", "Hosting, monitoreo, fixes de bugs"],
    ["Estándar", "199 – 349", "499 – 899", "129 – 249", "+ bolsa de horas para cambios menores"],
    ["Premium", "A medida", "A medida", "A medida", "+ SLA, horas dedicadas, prioridad"],
]
for row_vals in mant:
    data_row(ws, r, row_vals); r += 1

# column widths
widths = [40, 18, 18, 18, 34]
for i, w in enumerate(widths):
    ws.column_dimensions[get_column_letter(1 + i)].width = w

# ════════════════════════════════════════════════════════
# SHEET 2 — CALCULADORA (fórmulas vivas)
# ════════════════════════════════════════════════════════
cw = wb.create_sheet("Calculadora")
cw.sheet_view.showGridLines = False
cw["A1"] = "Calculadora de cotización"
cw["A1"].font = title_font
cw["A2"] = "Editá solo las celdas amarillas. El total se calcula solo."
cw["A2"].font = sub_font

# region -> hourly rate helper table (hidden-ish, lower area)
cw["E10"] = "Tabla de tarifas (no editar)"
cw["E10"].font = Font(bold=True, color=GREY, size=9)
rates = [["España", 55], ["EEUU", 90], ["LATAM", 35]]
for i, (reg, rate) in enumerate(rates):
    cw.cell(row=11 + i, column=5, value=reg).font = Font(size=9)
    cw.cell(row=11 + i, column=6, value=rate).font = Font(size=9)

def field(cw, row, label, fill=None, value=None):
    lc = cw.cell(row=row, column=1, value=label)
    lc.font = Font(bold=True, size=11)
    lc.alignment = left
    vc = cw.cell(row=row, column=2, value=value)
    vc.border = border
    vc.alignment = right
    if fill:
        vc.fill = fill
    return vc

field(cw, 4, "Región", input_fill, "España")
field(cw, 5, "Horas estimadas", input_fill, 80)
field(cw, 6, "Complejidad (1.0 / 1.3 / 1.6)", input_fill, 1.3)
field(cw, 7, "Costes de terceros (APIs, licencias)", input_fill, 0)

field(cw, 9, "Tarifa/hora (auto)").value = "=VLOOKUP(B4,E11:F13,2,FALSE)"
field(cw, 10, "Subtotal trabajo").value = "=B5*B9*B6"
field(cw, 11, "Colchón imprevistos (15%)").value = "=B10*0.15"
field(cw, 12, "Costes de terceros").value = "=B7"

total = field(cw, 13, "TOTAL PROYECTO", result_fill)
total.value = "=B10+B11+B12"
total.font = Font(bold=True, size=13, color="0A7D3F")
cw["A13"].font = Font(bold=True, size=13)

mant = field(cw, 14, "Mantenimiento sugerido (€/mes ≈17,5%/año)", result_fill)
mant.value = "=B13*0.175/12"
mant.font = Font(bold=True, size=11, color="0A7D3F")

cw["A16"] = "Fórmula: Total = (Horas × Tarifa × Complejidad) + 15% colchón + terceros"
cw["A16"].font = Font(italic=True, color=GREY, size=9)
cw["A17"] = "Cobrá por VALOR cuando puedas: el precio-hora es tu piso, no tu techo."
cw["A17"].font = Font(italic=True, color=GREY, size=9)

dv_region = DataValidation(type="list", formula1='"España,EEUU,LATAM"', allow_blank=False)
cw.add_data_validation(dv_region)
dv_region.add(cw["B4"])
dv_cx = DataValidation(type="list", formula1='"1,1.3,1.6"', allow_blank=False)
cw.add_data_validation(dv_cx)
dv_cx.add(cw["B6"])

cw.column_dimensions["A"].width = 40
cw.column_dimensions["B"].width = 16
cw.column_dimensions["E"].width = 12
cw.column_dimensions["F"].width = 8

# ════════════════════════════════════════════════════════
# SHEET 3 — MERCADO Y FUENTES
# ════════════════════════════════════════════════════════
mk = wb.create_sheet("Mercado y fuentes")
mk.sheet_view.showGridLines = False
mk["A1"] = "Referencias de mercado (2025)"
mk["A1"].font = title_font

header_row(mk, 3, ["Concepto", "España (EUR)", "EEUU (USD)", "LATAM (USD)"])
mref = [
    ["Hora freelance senior", "35 – 90", "50 – 300", "20 – 55"],
    ["Hora agencia", "80 – 120", "90 – 250", "30 – 80"],
    ["Hora especialista IA/ML", "~135", "100 – 200", "40 – 90"],
    ["MVP", "6.000 – 20.000", "15.000 – 50.000", "3.500 – 15.000"],
    ["Chatbot IA WhatsApp — setup", "2.000 – 5.000", "2.000 – 10.000", "450 – 1.500"],
    ["Chatbot IA WhatsApp — mensual", "49 – 500", "30 – 800", "110 – 275"],
    ["Mantenimiento", "15–20%/año", "15–20%/año", "15–20%/año"],
]
rr = 4
for row_vals in mref:
    data_row(mk, rr, row_vals); rr += 1

rr += 1
mk.cell(row=rr, column=1, value="Fuentes").font = Font(bold=True, color=GREEN, size=12)
rr += 1
sources = [
    "España — Armadillo Amarillo / CalcuTech / Hiberus / ABAMobile / Cronomia / Agencia IA Solutions",
    "EEUU — FullStack Labs / Arc.dev / Spdload / Fuzen / Tidio / AiSensy",
    "LATAM — Grupo Vansur (AR) / SODI / ToGrow / MY Tech Solutions (CO) / Simplixy (MX) / Kosmo (MX)",
    "Nota Meta: desde jul-2025 WhatsApp cobra por mensaje entregado (no por conversación de 24h).",
]
for s in sources:
    c = mk.cell(row=rr, column=1, value="• " + s)
    c.font = Font(size=9, color="555555")
    c.alignment = left
    mk.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=4)
    rr += 1

for i, w in enumerate([34, 20, 20, 20]):
    mk.column_dimensions[get_column_letter(1 + i)].width = w

# ════════════════════════════════════════════════════════
# SHEET 4 — ESTIMADOR DE ALCANCE (horas -> precio 3 regiones)
# ════════════════════════════════════════════════════════
es = wb.create_sheet("Estimador de alcance")
es.sheet_view.showGridLines = False
es["A1"] = "Estimador de alcance → precio"
es["A1"].font = title_font
es["A2"] = "Desglosá el proyecto en funcionalidades, asigná tamaño (S/M/L/XL) y el precio cae solo. Editá amarillo."
es["A2"].font = sub_font

# size -> hours reference table (editable)
es["H4"] = "Tamaño → horas (editable)"
es["H4"].font = Font(bold=True, color=GREY, size=9)
sizes = [["S", 8, "Una pantalla / CRUD simple"],
         ["M", 20, "Módulo con lógica"],
         ["L", 40, "Módulo complejo / integración"],
         ["XL", 80, "Subsistema completo"]]
for i, (s, h, desc) in enumerate(sizes):
    es.cell(row=5 + i, column=8, value=s).font = Font(size=9, bold=True)
    c = es.cell(row=5 + i, column=9, value=h); c.font = Font(size=9); c.fill = input_fill; c.border = border
    es.cell(row=5 + i, column=10, value=desc).font = Font(size=9, color=GREY)
# ── Catálogo de casos de uso por tamaño (lo que ofrece la web) ──
es.cell(row=10, column=8, value="Catálogo de casos de uso (ordenado por tamaño)").font = Font(bold=True, color=GREEN, size=11)
for i, h in enumerate(["Tamaño", "Horas", "Caso de uso"]):
    hc = es.cell(row=11, column=8 + i, value=h)
    hc.fill = dark_fill if i == 0 else green_fill
    hc.font = hdr_font; hc.alignment = center; hc.border = border
catalog = [
    # S — 8h
    ("S", 8, "Landing / página informativa"),
    ("S", 8, "Formulario de contacto + captura de lead"),
    ("S", 8, "Multi-idioma (i18n)"),
    ("S", 8, "Notificaciones por email (transaccionales)"),
    ("S", 8, "CRUD de entidad simple"),
    ("S", 8, "Listado / catálogo de productos"),
    # M — 20h
    ("M", 20, "Login + registro + roles"),
    ("M", 20, "Setup de proyecto + deploy / CI"),
    ("M", 20, "Panel de usuario / perfil"),
    ("M", 20, "CRUD con relaciones y validaciones"),
    ("M", 20, "Búsqueda, filtros y paginación"),
    ("M", 20, "Carga y gestión de archivos (CV, docs)"),
    ("M", 20, "Flujo de estados / aprobación simple"),
    # L — 40h
    ("L", 40, "Dashboard con reportes y métricas"),
    ("L", 40, "Integración con API externa"),
    ("L", 40, "Pasarela de pago"),
    ("L", 40, "PWA / funcionalidad offline-first"),
    ("L", 40, "Sincronización en tiempo real"),
    ("L", 40, "Integración WhatsApp Business API"),
    ("L", 40, "Motor de cálculo automático (raciones, precios)"),
    # XL — 80h
    ("XL", 80, "Agente IA WhatsApp (contacta, califica, agenda)"),
    ("XL", 80, "CRM completo (pipeline + asignación de leads)"),
    ("XL", 80, "Panel de administración multi-módulo"),
    ("XL", 80, "Sistema multi-sede / multi-centro"),
    ("XL", 80, "MVP funcional completo (4–8 semanas)"),
    ("XL", 80, "Transformación digital (diagnóstico + roadmap)"),
]
size_fill = {
    "S": PatternFill("solid", fgColor="EAF7EF"),
    "M": PatternFill("solid", fgColor="DDEFFF"),
    "L": PatternFill("solid", fgColor="FFF1DD"),
    "XL": PatternFill("solid", fgColor="F3E8FF"),
}
for i, (s, h, case) in enumerate(catalog):
    row = 12 + i
    sc = es.cell(row=row, column=8, value=s)
    sc.font = Font(bold=True, size=9); sc.alignment = center; sc.border = border; sc.fill = size_fill[s]
    hc = es.cell(row=row, column=9, value=h)
    hc.font = Font(size=9); hc.alignment = center; hc.border = border; hc.fill = size_fill[s]
    cc = es.cell(row=row, column=10, value=case)
    cc.font = Font(size=9, color="333333"); cc.alignment = left; cc.border = border; cc.fill = size_fill[s]

# feature breakdown table
header_row(es, 4, ["Funcionalidad", "Tamaño", "Horas"])
es.column_dimensions["A"].width = 38
es.column_dimensions["B"].width = 12
es.column_dimensions["C"].width = 10
for col in ("E", "F", "G"):
    es.column_dimensions[col].width = 16
es.column_dimensions["H"].width = 10
es.column_dimensions["I"].width = 9
es.column_dimensions["J"].width = 46

example = ["Login / registro / roles", "Setup + deploy", "Panel principal (dashboard)",
           "CRUD entidad principal", "Integración WhatsApp + IA"]
FIRST, LAST = 5, 19
dv_size = DataValidation(type="list", formula1='"S,M,L,XL"', allow_blank=True)
es.add_data_validation(dv_size)
for row in range(FIRST, LAST + 1):
    fa = es.cell(row=row, column=1, value=example[row - FIRST] if row - FIRST < len(example) else None)
    fa.border = border; fa.alignment = left; fa.fill = input_fill
    sb = es.cell(row=row, column=2, value="M" if row - FIRST < len(example) else None)
    sb.border = border; sb.alignment = center; sb.fill = input_fill
    dv_size.add(sb)
    hc = es.cell(row=row, column=3, value=f"=IFERROR(VLOOKUP(B{row},$H$5:$I$8,2,FALSE),0)")
    hc.border = border; hc.alignment = right

es.cell(row=20, column=1, value="Subtotal funcionalidades").font = Font(bold=True)
st = es.cell(row=20, column=3, value=f"=SUM(C{FIRST}:C{LAST})")
st.font = Font(bold=True); st.border = border; st.alignment = right

def pct_row(row, label, default):
    es.cell(row=row, column=1, value=label).font = Font(size=10)
    p = es.cell(row=row, column=2, value=default)
    p.fill = input_fill; p.border = border; p.alignment = right
    p.number_format = "0%"
    h = es.cell(row=row, column=3, value=f"=$C$20*B{row}")
    h.border = border; h.alignment = right
    return row

pct_row(22, "Gestión de proyecto (%)", 0.12)
pct_row(23, "Testing / QA (%)", 0.15)
pct_row(24, "Contingencia (%)", 0.10)

es.cell(row=26, column=1, value="TOTAL HORAS").font = Font(bold=True, size=13)
th = es.cell(row=26, column=3, value="=C20+C22+C23+C24")
th.font = Font(bold=True, size=13, color="0A7D3F"); th.fill = result_fill
th.border = border; th.alignment = right

# region pricing block
es.cell(row=28, column=1, value="PRECIO POR REGIÓN").font = Font(bold=True, color=GREEN, size=13)
header_row(es, 29, ["Región", "Tarifa/hora", "Precio proyecto", "Mantenim./mes"])
regions = [("España (EUR)", 55), ("EEUU (USD)", 90), ("LATAM (USD)", 35)]
for i, (name, rate) in enumerate(regions):
    rr = 30 + i
    es.cell(row=rr, column=1, value=name).font = Font(bold=True, size=10)
    es.cell(row=rr, column=1).border = border
    tc = es.cell(row=rr, column=2, value=rate)
    tc.fill = input_fill; tc.border = border; tc.alignment = right
    pc = es.cell(row=rr, column=3, value=f"=$C$26*B{rr}")
    pc.font = Font(bold=True, color="0A7D3F"); pc.fill = result_fill
    pc.border = border; pc.alignment = right
    mc = es.cell(row=rr, column=4, value=f"=C{rr}*0.175/12")
    mc.fill = result_fill; mc.border = border; mc.alignment = right

es.cell(row=34, column=1, value="La tarifa/hora es editable: bajala para clientes de entrada, subila cuando tengas portfolio.").font = Font(italic=True, color=GREY, size=9)
es.cell(row=35, column=1, value="El precio sale del trabajo real → es defendible ante el cliente. Cobrá por VALOR cuando el impacto lo justifique.").font = Font(italic=True, color=GREY, size=9)

import sys
out = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\torre\Documents\nexdevp\nexdevp-precios.xlsx"
wb.save(out)
print("saved:", out)
